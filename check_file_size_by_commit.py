#!/usr/bin/env python3
import argparse
import requests
import os
import pandas as pd
from tqdm import tqdm
import time
import json
from pathlib import Path
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timezone
import logging
from typing import List, Dict, Any, Optional, Callable
import re
import colorama
from colorama import Fore, Style
from tqdm.auto import tqdm
import sys
import signal
import urllib3
import warnings
import urllib.parse
import subprocess
import zipfile
import shutil

# Suppress SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize colorama
colorama.init()

def signal_handler(signum, frame):
    print(f"\n{Fore.YELLOW}⚠️ Proses dibatalkan oleh user{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}ℹ️ Data yang sudah diproses akan disimpan{Style.RESET_ALL}")
    sys.exit(0)

# Register signal handler
signal.signal(signal.SIGINT, signal_handler)

# Custom formatter untuk menambahkan warna
class ColoredFormatter(logging.Formatter):
    def format(self, record):
        if record.levelno >= logging.ERROR:
            record.msg = f"{Fore.RED}{record.msg}{Style.RESET_ALL}"
        return super().format(record)

# Set formatter
handler = logging.StreamHandler()
handler.setFormatter(ColoredFormatter('%(asctime)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

class ErrorCollector:
    def __init__(self):
        self.errors = []
    
    def add_error(self, error_msg: str):
        self.errors.append(error_msg)
    
    def has_errors(self) -> bool:
        return len(self.errors) > 0
    
    def print_errors(self):
        if self.errors:
            print(f"\n{Fore.RED}=== Error Summary ==={Style.RESET_ALL}")
            for i, error in enumerate(self.errors, 1):
                print(f"{Fore.RED}{i}. {error}{Style.RESET_ALL}")
            print(f"{Fore.RED}==================={Style.RESET_ALL}\n")

class GitLabAPI:
    def __init__(self, gitlab_url: str, token: str, project_id: str, verify_ssl: bool = True):
        self.api_url = gitlab_url.rstrip('/') + '/api/v4'
        self.headers = {'PRIVATE-TOKEN': token}
        self.project_id = project_id
        self.cache_dir = Path('.gitlab_cache')
        self.cache_dir.mkdir(exist_ok=True)
        self.verify_ssl = verify_ssl
        self.error_collector = ErrorCollector()
        
    def _get_cache_path(self, endpoint: str, params: Dict) -> Path:
        cache_key = f"{endpoint}_{hash(json.dumps(params, sort_keys=True))}.json"
        return self.cache_dir / cache_key
    
    def _get_cached_data(self, cache_path: Path) -> Optional[List[Dict]]:
        if cache_path.exists():
            with open(cache_path, 'r') as f:
                return json.load(f)
        return None
    
    def _save_to_cache(self, cache_path: Path, data: List[Dict]):
        with open(cache_path, 'w') as f:
            json.dump(data, f)
    
    def get_merge_requests(self, target_branch: str, states: List[str], commit_message: Optional[str] = None) -> List[Dict]:
        all_mrs = []
        for state in states:
            page = 1
            while True:
                params = {
                    "state": state,
                    "target_branch": target_branch,
                    "per_page": 100,
                    "page": page,
                    "order_by": "updated_at",
                    "sort": "desc"
                }
                
                cache_path = self._get_cache_path('merge_requests', params)
                cached_data = self._get_cached_data(cache_path)
                
                if cached_data:
                    mrs = cached_data
                else:
                    try:
                        url = f"{self.api_url}/projects/{self.project_id}/merge_requests"
                        r = requests.get(url, headers=self.headers, params=params, verify=self.verify_ssl)
                        r.raise_for_status()
                        mrs = r.json()
                        
                        if not mrs:
                            break
                            
                        self._save_to_cache(cache_path, mrs)
                        
                        # Rate limiting
                        time.sleep(0.5)
                        
                    except requests.exceptions.RequestException as e:
                        logger.error(f"Error fetching merge requests: {e}")
                        break
                
                # Filter MRs based on commit message if provided
                if commit_message:
                    filtered_mrs = []
                    for mr in mrs:
                        try:
                            # Get commits for this MR
                            commits_url = f"{self.api_url}/projects/{self.project_id}/merge_requests/{mr['iid']}/commits"
                            commits_r = requests.get(commits_url, headers=self.headers, verify=self.verify_ssl)
                            commits_r.raise_for_status()
                            commits = commits_r.json()
                            
                            # Check if any commit message contains the search string
                            if any(commit_message.lower() in commit['title'].lower() for commit in commits):
                                filtered_mrs.append(mr)
                            
                            # Rate limiting
                            time.sleep(0.5)
                            
                        except requests.exceptions.RequestException as e:
                            logger.error(f"Error fetching commits for MR {mr['iid']}: {e}")
                            continue
                    
                    mrs = filtered_mrs
                
                all_mrs.extend(mrs)
                page += 1
                
                if not mrs:
                    break
                    
        return all_mrs
    
    def get_mr_changes(self, mr_iid: int) -> List[Dict]:
        cache_path = self._get_cache_path('mr_changes', {'mr_iid': mr_iid})
        cached_data = self._get_cached_data(cache_path)
        
        if cached_data:
            return cached_data
            
        try:
            print(f"\n🔍 Checking MR #{mr_iid}...")
            url = f"{self.api_url}/projects/{self.project_id}/merge_requests/{mr_iid}/changes"
            r = requests.get(url, headers=self.headers, verify=self.verify_ssl)
            r.raise_for_status()
            changes = r.json()["changes"]
            
            # Tambahkan informasi ukuran file untuk setiap perubahan
            for change in changes:
                if 'new_path' in change:
                    try:
                        # Gunakan endpoint yang benar untuk mendapatkan file
                        file_url = f"{self.api_url}/projects/{self.project_id}/repository/files/{urllib.parse.quote(change['new_path'], safe='')}/blob"
                        file_params = {'ref': change.get('new_sha', '')}
                        print(f"\nChecking file: {change['new_path']}")
                        print(f"URL: {file_url}")
                        print(f"Params: {file_params}")
                        
                        file_r = requests.get(file_url, headers=self.headers, params=file_params, verify=self.verify_ssl)
                        
                        if file_r.status_code == 200:
                            file_data = file_r.json()
                            # Ukuran file ada di response
                            if 'size' in file_data:
                                size_bytes = file_data['size']
                                size_kb = round(size_bytes / 1024, 2)
                                change['size_kb'] = size_kb
                                print(f"File size: {size_kb:.2f}KB ({size_kb/1024:.2f}MB)")
                            else:
                                change['size_kb'] = None
                                print(f"⚠️ No size information for {change['new_path']}")
                        else:
                            # Jika file tidak ditemukan, coba dapatkan dari diff
                            if 'diff' in change:
                                # Hitung ukuran dari diff
                                size_bytes = len(change['diff'].encode('utf-8'))
                                size_kb = round(size_bytes / 1024, 2)
                                change['size_kb'] = size_kb
                                print(f"File size from diff: {size_kb:.2f}KB")
                            else:
                                change['size_kb'] = None
                                print(f"⚠️ Could not get file size for {change['new_path']}")
                                print(f"Status code: {file_r.status_code}")
                    except Exception as e:
                        self.error_collector.add_error(f"Error getting file size for {change['new_path']}: {e}")
                        change['size_kb'] = None
            
            self._save_to_cache(cache_path, changes)
            return changes
        except requests.exceptions.RequestException as e:
            self.error_collector.add_error(f"Error fetching MR changes for MR {mr_iid}: {e}")
            return []

    def _encode_branch_name(self, branch: str) -> str:
        """Encode branch name for URL"""
        return branch.replace('/', '%2F')

    def get_branch_creation_date(self, branch: str) -> Optional[datetime]:
        """Get the creation date of a branch"""
        try:
            encoded_branch = self._encode_branch_name(branch)
            url = f"{self.api_url}/projects/{self.project_id}/repository/branches/{encoded_branch}"
            r = requests.get(url, headers=self.headers, verify=self.verify_ssl)
            r.raise_for_status()
            branch_info = r.json()
            return datetime.fromisoformat(branch_info['commit']['created_at'].replace('Z', '+00:00'))
        except requests.exceptions.RequestException as e:
            self.error_collector.add_error(f"Error fetching branch creation date: {e}")
            return None

    def get_commit_count(self, branch: str, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None) -> int:
        """Get total number of commits in a date range"""
        try:
            url = f"{self.api_url}/projects/{self.project_id}/repository/commits"
            params = {
                "ref_name": branch,
                "per_page": 1,  # Hanya butuh 1 commit untuk mendapatkan total
                "page": 1,
                "order": "default"
            }
            
            # Tambahkan filter tanggal
            if start_date:
                params["after"] = start_date.isoformat()
            if end_date:
                params["before"] = end_date.isoformat()
            
            r = requests.get(url, headers=self.headers, params=params, verify=self.verify_ssl)
            r.raise_for_status()
            
            # Total commit ada di header X-Total
            if 'X-Total' in r.headers:
                return int(r.headers['X-Total'])
            return 0
            
        except requests.exceptions.RequestException as e:
            self.error_collector.add_error(f"Error getting commit count: {e}")
            return 0

    def get_commits_by_branch(self, branch: str, after_date: Optional[datetime] = None, limit: Optional[int] = None, 
                            start_date: Optional[datetime] = None, end_date: Optional[datetime] = None,
                            callback: Optional[Callable] = None) -> List[Dict]:
        """Get commits from a specific branch and process them immediately"""
        print(f"\n🔍 Checking branch: {branch}")
        
        # First, verify branch exists
        try:
            encoded_branch = self._encode_branch_name(branch)
            url = f"{self.api_url}/projects/{self.project_id}/repository/branches/{encoded_branch}"
            r = requests.get(url, headers=self.headers, verify=self.verify_ssl)
            r.raise_for_status()
            branch_info = r.json()
            print(f"\n✅ Branch found:")
            print(f"Name: {branch_info['name']}")
            print(f"Commit: {branch_info['commit']['id']}")
            print(f"Created at: {branch_info['commit']['created_at']}")
        except requests.exceptions.RequestException as e:
            print(f"\n❌ Error accessing branch: {e}")
            return
        
        # Now get commits with date filter
        print("\n📥 Fetching commits...")
        url = f"{self.api_url}/projects/{self.project_id}/repository/commits"
        
        # Add date filters
        params = {
            "ref_name": branch,
            "per_page": 100,  # Get maximum commits per page
            "page": 1,
            "order": "default"
        }
        
        if start_date:
            # Format: YYYY-MM-DD
            params["after"] = start_date.strftime("%Y-%m-%d")
            print(f"Start date: {params['after']}")
        if end_date:
            # Format: YYYY-MM-DD
            params["before"] = end_date.strftime("%Y-%m-%d")
            print(f"End date: {params['before']}")
        
        try:
            all_commits = []
            page = 1
            total_commits = 0
            
            while True:
                params["page"] = page
                r = requests.get(url, headers=self.headers, params=params, verify=self.verify_ssl)
                r.raise_for_status()
                commits = r.json()
                
                if not commits:
                    break
                
                # Filter commits by date
                for commit in commits:
                    commit_date = datetime.fromisoformat(commit['created_at'].replace('Z', '+00:00'))
                    # Skip if commit is outside date range
                    if start_date and commit_date.date() < start_date.date():
                        continue
                    if end_date and commit_date.date() > end_date.date():
                        continue
                    all_commits.append(commit)
                
                # If we've gone past our date range, stop fetching
                if commits and start_date:
                    last_commit_date = datetime.fromisoformat(commits[-1]['created_at'].replace('Z', '+00:00'))
                    if last_commit_date.date() < start_date.date():
                        break
                
                total_commits += len(commits)
                print(f"\rFetched {total_commits} commits...", end="")
                
                if len(commits) < 100:  # Last page
                    break
                    
                page += 1
                time.sleep(0.5)  # Rate limiting
            
            print(f"\n\n✅ Found {len(all_commits)} commits in date range")
            
            if all_commits:
                print("\n📋 Commits in date range:")
                for commit in all_commits[:5]:  # Show first 5 commits
                    commit_date = datetime.fromisoformat(commit['created_at'].replace('Z', '+00:00'))
                    print(f"ID: {commit['id'][:8]} | Date: {commit_date.strftime('%Y-%m-%d')} | Title: {commit['title']}")
                
                # Process filtered commits
                processed = 0
                for commit in all_commits:
                    if callback:
                        callback(commit)
                    processed += 1
                    if limit and processed >= limit:
                        break
                
                print(f"\n✅ Processed {processed} commits")
            else:
                print("\n❌ No commits found in the specified date range")
            
        except requests.exceptions.RequestException as e:
            print(f"\n❌ Error fetching commits: {e}")
            return

    def get_commit_changes(self, commit_id: str) -> List[Dict]:
        """Get changes for a specific commit"""
        cache_path = self._get_cache_path('commit_changes', {'commit_id': commit_id})
        cached_data = self._get_cached_data(cache_path)
        
        if cached_data:
            return cached_data
            
        try:
            url = f"{self.api_url}/projects/{self.project_id}/repository/commits/{commit_id}/diff"
            r = requests.get(url, headers=self.headers, verify=self.verify_ssl)
            r.raise_for_status()
            changes = r.json()
            
            # Tambahkan informasi ukuran file untuk setiap perubahan
            for change in changes:
                if 'new_path' in change:
                    try:
                        # Gunakan endpoint yang benar untuk mendapatkan file
                        file_url = f"{self.api_url}/projects/{self.project_id}/repository/files/{urllib.parse.quote(change['new_path'], safe='')}/blob"
                        file_params = {'ref': commit_id}
                        file_r = requests.get(file_url, headers=self.headers, params=file_params, verify=self.verify_ssl)
                        
                        if file_r.status_code == 200:
                            file_data = file_r.json()
                            # Ukuran file ada di response
                            if 'size' in file_data:
                                size_bytes = file_data['size']
                                size_kb = round(size_bytes / 1024, 2)
                                change['size_kb'] = size_kb
                                # Log file besar
                                if size_kb > 1000:  # File lebih dari 1MB
                                    print(f"\n⚠️ Large file detected in commit {commit_id[:8]}:")
                                    print(f"File: {change['new_path']}")
                                    print(f"Size: {size_kb:.2f}KB ({size_kb/1024:.2f}MB)")
                            else:
                                change['size_kb'] = None
                                print(f"\n⚠️ No size information for {change['new_path']} in commit {commit_id[:8]}")
                        else:
                            # Jika file tidak ditemukan, coba dapatkan dari diff
                            if 'diff' in change:
                                # Hitung ukuran dari diff
                                size_bytes = len(change['diff'].encode('utf-8'))
                                size_kb = round(size_bytes / 1024, 2)
                                change['size_kb'] = size_kb
                            else:
                                change['size_kb'] = None
                                print(f"\n⚠️ Could not get file size for {change['new_path']} in commit {commit_id[:8]}")
                                print(f"Status code: {file_r.status_code}")
                    except Exception as e:
                        self.error_collector.add_error(f"Error getting file size for {change['new_path']}: {e}")
                        change['size_kb'] = None
            
            self._save_to_cache(cache_path, changes)
            return changes
        except requests.exceptions.RequestException as e:
            self.error_collector.add_error(f"Error fetching changes for commit {commit_id}: {e}")
            return []

    def check_branch_and_commits(self, branch: str):
        """Check branch existence and get some commits without date filter"""
        try:
            # Check branch existence
            encoded_branch = self._encode_branch_name(branch)
            url = f"{self.api_url}/projects/{self.project_id}/repository/branches/{encoded_branch}"
            r = requests.get(url, headers=self.headers, verify=self.verify_ssl)
            r.raise_for_status()
            branch_info = r.json()
            print(f"\n📌 Branch info:")
            print(f"Name: {branch_info['name']}")
            print(f"Commit: {branch_info['commit']['id']}")
            print(f"Created at: {branch_info['commit']['created_at']}")
            
            # Get some commits without date filter
            url = f"{self.api_url}/projects/{self.project_id}/repository/commits"
            params = {
                "ref_name": branch,
                "per_page": 5,  # Get only 5 commits for sample
                "page": 1
            }
            
            r = requests.get(url, headers=self.headers, params=params, verify=self.verify_ssl)
            r.raise_for_status()
            commits = r.json()
            
            print(f"\n📦 Sample commits:")
            for commit in commits:
                print(f"ID: {commit['id'][:8]} | Date: {commit['created_at']} | Title: {commit['title']}")
            
            return True
            
        except requests.exceptions.RequestException as e:
            print(f"\n❌ Error checking branch: {e}")
            return False

class FileAnalyzer:
    def __init__(self, file_patterns: Optional[List[str]] = None, 
                 min_size_kb: Optional[float] = None,
                 max_size_kb: Optional[float] = None):
        self.file_patterns = [re.compile(pattern) for pattern in (file_patterns or [])]
        self.min_size_kb = min_size_kb
        self.max_size_kb = max_size_kb
        self.error_collector = ErrorCollector()
    
    def should_analyze_file(self, filepath: str, size_kb: float) -> bool:
        if self.file_patterns and not any(pattern.search(filepath) for pattern in self.file_patterns):
            return False
        if self.min_size_kb and size_kb < self.min_size_kb:
            return False
        if self.max_size_kb and size_kb > self.max_size_kb:
            return False
        return True
    
    def get_file_size(self, filepath: str) -> Optional[float]:
        try:
            size_bytes = os.path.getsize(filepath)
            size_kb = round(size_bytes / 1024, 2)
            return size_kb if self.should_analyze_file(filepath, size_kb) else None
        except Exception as e:
            self.error_collector.add_error(f"Error getting file size for {filepath}: {e}")
            return None

# Tambahkan di bagian atas file
NON_STANDARD_EXTS = ['.apk', '.aab', '.so', '.jar', '.dex', '.class', '.aar']

def is_non_standard(filepath):
    return any(filepath.lower().endswith(ext) for ext in NON_STANDARD_EXTS)

class ReportGenerator:
    def __init__(self, output_excel: str):
        self.output_excel = output_excel
        self.data = []
        self.commit_summary = {}  # Untuk menyimpan ringkasan per commit
        self.error_collector = ErrorCollector()
    
    def add_data(self, mr_title: str, mr_id: int, mr_state: str, 
                file_path: str, file_size: Optional[float], non_standard: Optional[bool] = None):
        if non_standard is None:
            non_standard = is_non_standard(file_path)
        self.data.append({
            "MR Title": mr_title,
            "MR ID": mr_id,
            "MR State": mr_state,
            "File": file_path,
            "File Size (KB)": file_size,
            "NonStandard": non_standard
        })
        
        # Update ringkasan commit
        if mr_id not in self.commit_summary:
            self.commit_summary[mr_id] = {
                'title': mr_title,
                'total_size': 0,
                'file_count': 0,
                'files': []
            }
        
        if file_size is not None:
            self.commit_summary[mr_id]['total_size'] += file_size
            self.commit_summary[mr_id]['file_count'] += 1
            self.commit_summary[mr_id]['files'].append({
                'path': file_path,
                'size': file_size
            })
    
    def generate_excel(self):
        try:
            # Buat DataFrame dan urutkan berdasarkan ukuran file
            df = pd.DataFrame(self.data)
            if not df.empty and 'File Size (KB)' in df.columns:
                df = df.sort_values('File Size (KB)', ascending=False)
            
            # Simpan ke Excel
            with pd.ExcelWriter(self.output_excel, engine='openpyxl') as writer:
                # Sheet untuk data file
                df.to_excel(writer, sheet_name='File Analysis', index=False)
                
                # Sheet untuk ringkasan commit
                if self.commit_summary:
                    summary_data = []
                    for commit_id, info in self.commit_summary.items():
                        summary_data.append({
                            'Commit ID': commit_id,
                            'Title': info['title'],
                            'Total Size (KB)': round(info['total_size'], 2),
                            'File Count': info['file_count']
                        })
                    
                    summary_df = pd.DataFrame(summary_data)
                    summary_df = summary_df.sort_values('Total Size (KB)', ascending=False)
                    summary_df.to_excel(writer, sheet_name='Commit Summary', index=False)
                    
                    # Tambahkan sheet untuk detail file per commit
                    for commit_id, info in self.commit_summary.items():
                        files_df = pd.DataFrame(info['files'])
                        files_df = files_df.sort_values('size', ascending=False)
                        sheet_name = f'Commit_{commit_id[:8]}'
                        files_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"✅ Data disimpan ke file: {self.output_excel}")
            
            # Tampilkan ringkasan commit terbesar
            if self.commit_summary:
                print("\n📊 Top 5 Commits by Size:")
                print("=" * 80)
                print(f"{'Commit ID':<10} {'Total Size (KB)':<15} {'File Count':<12} Title")
                print("-" * 80)
                
                sorted_commits = sorted(
                    self.commit_summary.items(),
                    key=lambda x: x[1]['total_size'],
                    reverse=True
                )[:5]
                
                for commit_id, info in sorted_commits:
                    print(f"{commit_id[:8]:<10} {info['total_size']:<15.2f} {info['file_count']:<12} {info['title']}")
                print("=" * 80)
            
            return df
        except Exception as e:
            self.error_collector.add_error(f"Error generating Excel file: {e}")
            return pd.DataFrame()
    
    def generate_plots(self, df: pd.DataFrame):
        if df.empty or 'File Size (KB)' not in df.columns or df['File Size (KB)'].dropna().empty:
            logger.warning("Data kosong, tidak ada plot yang dibuat.")
            return
            
        try:
            # Create plots directory
            plots_dir = Path('plots')
            plots_dir.mkdir(exist_ok=True)
            
            # Plot 1: File Size Distribution
            plt.figure(figsize=(12, 6))
            sns.histplot(data=df, x='File Size (KB)', bins=30)
            plt.title('Distribusi Ukuran File')
            plt.savefig(plots_dir / 'file_size_distribution.png')
            plt.close()
            
            # Plot 2: Top 10 Largest Files
            plt.figure(figsize=(12, 6))
            top_files = df.nlargest(10, 'File Size (KB)')
            sns.barplot(data=top_files, x='File Size (KB)', y='File')
            plt.title('10 File Terbesar')
            plt.tight_layout()
            plt.savefig(plots_dir / 'top_10_largest_files.png')
            plt.close()
            
            # Plot 3: Commit Size Distribution
            if self.commit_summary:
                commit_sizes = [info['total_size'] for info in self.commit_summary.values()]
                plt.figure(figsize=(12, 6))
                sns.histplot(data=commit_sizes, bins=30)
                plt.title('Distribusi Ukuran Commit')
                plt.xlabel('Total Size (KB)')
                plt.savefig(plots_dir / 'commit_size_distribution.png')
                plt.close()
            
            logger.info("✅ Visualisasi disimpan di folder 'plots'")
        except Exception as e:
            self.error_collector.add_error(f"Error generating plots: {e}")

def get_file_size_in_commit(repo_path, commit_sha, file_path):
    """Get file size in KB for a file at a specific commit using git ls-tree (no checkout needed)"""
    cmd = [
        "git", "-C", repo_path, "ls-tree", "-l", commit_sha, file_path
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.stdout:
        parts = result.stdout.strip().split()
        if len(parts) >= 4:
            try:
                return int(parts[3]) / 1024  # size in KB
            except Exception:
                return None
    return None

def validate_file_size(file_path, size_mb):
    file_lower = file_path.lower()
    if file_lower.endswith(('.xml',)):
        return "OK" if size_mb is not None and size_mb <= 0.02 else "OVERSIZE"  # 20 KB
    if file_lower.endswith(('.png', '.jpg', '.jpeg')):
        return "OK" if size_mb is not None and size_mb <= 0.05 else "OVERSIZE"  # 50 KB
    if file_lower.endswith(('.webp',)):
        return "OK" if size_mb is not None and size_mb <= 0.2 else "OVERSIZE"   # 200 KB
    if file_lower.endswith(('.ogg', '.aac')):
        return "OK" if size_mb is not None and size_mb <= 0.3 else "OVERSIZE"   # 300 KB
    if file_lower.endswith(('.mp4', '.mov', '.m4v')):
        return "OK" if size_mb is not None and size_mb <= 1 else "OVERSIZE"     # 1 MB
    if file_lower.endswith(('.json',)):
        return "OK" if size_mb is not None and size_mb <= 0.1 else "OVERSIZE"   # 100 KB
    if file_lower.endswith(('.ttf', '.otf')):
        return "OK" if size_mb is not None and size_mb <= 0.5 else "OVERSIZE"   # 500 KB
    if file_lower.endswith(('.so',)):
        return "OK" if size_mb is not None and size_mb <= 5 else "OVERSIZE"     # 5 MB
    if file_lower.endswith(('.dex',)):
        return "OK" if size_mb is not None and size_mb <= 10 else "OVERSIZE"    # 10 MB
    # Default: OK
    return "OK"

def suggest_optimization(file_path, size_mb):
    ext = file_path.lower()
    if ext.endswith(('.png', '.jpg', '.jpeg')):
        return "Kompres ke WebP, turunkan resolusi/quality"
    if ext.endswith('.webp'):
        return "Pastikan sudah lossy, cek resolusi"
    if ext.endswith(('.ogg', '.aac', '.mp3')):
        return "Turunkan bitrate, kompres audio"
    if ext.endswith('.mp4'):
        return "Turunkan resolusi/bitrate video"
    if ext.endswith(('.json', '.xml')):
        return "Pertimbangkan compress GZIP atau split"
    if ext.endswith(('.ttf', '.otf')):
        return "Pisahkan font ke modul terpisah"
    if ext.endswith(('.so', '.dex', '.aar', '.apk', '.jar')):
        return "File binary besar, audit manual kebutuhan file"
    return "Audit manual, cek kebutuhan file"

RELEVANT_FOLDERS = [
    'res', 'assets', 'jniLibs', 'lib', 'raw', 'fonts'
]
RELEVANT_EXTS = [
    '.png', '.jpg', '.jpeg', '.webp', '.xml', '.json', '.mp3', '.ogg', '.aac', '.ttf', '.otf', '.so', '.dex', '.aar', '.jar', '.apk', '.mp4'
]

def is_relevant_file(file_path):
    parts = file_path.split(os.sep)
    if not any(folder in parts for folder in RELEVANT_FOLDERS):
        return False
    if not any(file_path.lower().endswith(ext) for ext in RELEVANT_EXTS):
        return False
    return True

def analyze_local_snapshot(repo_path, output_excel="local_snapshot_report.xlsx", file_types=None):
    import os
    import pandas as pd
    from tqdm import tqdm
    file_list = []
    for root, dirs, files in os.walk(repo_path):
        if '.git' in root:
            continue
        for file in files:
            file_list.append(os.path.join(root, file))
    data = []
    for file_path in tqdm(file_list, desc="Snapshot HEAD", unit="file"):
        rel_path = os.path.relpath(file_path, repo_path)
        if not is_relevant_file(rel_path):
            continue
        # Filter by file_types if provided
        if file_types:
            if not any(rel_path.lower().endswith('.' + ext.strip().lower()) for ext in file_types):
                continue
        try:
            size_mb = round(os.path.getsize(file_path) / 1024 / 1024, 2)
        except Exception:
            size_mb = None
        data.append({
            "File": rel_path,
            "File Size (MB)": size_mb,
            "NonStandard": is_non_standard(file_path),
            "Validation": validate_file_size(file_path, size_mb)
        })
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(["File Size (MB)"], ascending=[False])
    # Sheet Optimization Candidates
    if not df.empty:
        opt_df = df[(df["Validation"] == "OVERSIZE") & (~df["File"].str.lower().str.endswith(tuple([".so", ".dex", ".aar", ".apk", ".jar"])))].copy()
        if not opt_df.empty:
            opt_df["Saran Optimasi"] = opt_df.apply(lambda row: suggest_optimization(row["File"], row["File Size (MB)"]), axis=1)
        else:
            opt_df = pd.DataFrame()
    else:
        opt_df = pd.DataFrame()
    return df, opt_df

def analyze_local_all_commits(repo_path, file_patterns=None):
    import subprocess
    import pandas as pd
    import re
    from tqdm import tqdm
    # Get all commits (not just linear)
    log_cmd = [
        "git", "-C", repo_path, "rev-list", "--all"
    ]
    result = subprocess.run(log_cmd, capture_output=True, text=True)
    commits = result.stdout.strip().splitlines()
    # Get commit info (date, title)
    commit_info = {}
    for sha in commits:
        info_cmd = [
            "git", "-C", repo_path, "show", "-s", "--format=%ad|%s", "--date=short", sha
        ]
        info_result = subprocess.run(info_cmd, capture_output=True, text=True)
        if '|' in info_result.stdout:
            date, title = info_result.stdout.strip().split('|', 1)
        else:
            date, title = '', ''
        commit_info[sha] = (date, title)
    data = []
    for sha in tqdm(commits, desc="Processing all commits", unit="commit"):
        diff_cmd = [
            "git", "-C", repo_path, "diff-tree", "--no-commit-id", "--name-only", "-r", "-m", "--root", sha
        ]
        diff_result = subprocess.run(diff_cmd, capture_output=True, text=True)
        files = diff_result.stdout.strip().splitlines()
        for file_path in files:
            if file_patterns and not any(re.search(p, file_path) for p in file_patterns):
                continue
            size_kb = get_file_size_in_commit(repo_path, sha, file_path)
            size_mb = round(size_kb / 1024, 2) if size_kb is not None else None
            validation = validate_file_size(file_path, size_mb)
            date, title = commit_info.get(sha, ('', ''))
            data.append({
                "Commit": sha[:8],
                "Date": date,
                "Commit Title": title,
                "File": file_path,
                "File Size (MB)": size_mb,
                "NonStandard": is_non_standard(file_path),
                "Validation": validation
            })
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(["File Size (MB)"], ascending=[False])
    return df

def get_current_branch(repo_path):
    import subprocess
    result = subprocess.run(
        ["git", "-C", repo_path, "rev-parse", "--abbrev-ref", "HEAD"],
        capture_output=True, text=True
    )
    return result.stdout.strip()

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--gitlab-url', help='GitLab URL (e.g., https://gitlab.com)')
    parser.add_argument('--token', help='GitLab private token')
    parser.add_argument('--project-id', help='GitLab project ID')
    parser.add_argument('--target-branch', help='Target branch (e.g., release/8.6.0/master)')
    parser.add_argument('--mr-state', default='opened', help='MR states (comma-separated: opened,merged,closed)')
    parser.add_argument('--output-excel', default='output.xlsx', help='Output Excel filename')
    parser.add_argument('--file-patterns', help='File patterns to include (comma-separated regex)')
    parser.add_argument('--min-size-kb', type=float, help='Minimum file size in KB to include')
    parser.add_argument('--max-size-kb', type=float, help='Maximum file size in KB to include')
    parser.add_argument('--commit-message', help='Filter MRs by commit message (case-insensitive)')
    parser.add_argument('--no-plot', action='store_true', help='Skip plotting')
    parser.add_argument('--no-verify-ssl', action='store_true', help='Disable SSL certificate verification')
    parser.add_argument('--analyze-branch', help='Analyze file sizes in a specific branch')
    parser.add_argument('--commit-title', help='Filter commits by title (case-insensitive)')
    parser.add_argument('--after-branch-creation', action='store_true', help='Only analyze commits created after the branch was created')
    parser.add_argument('--start-date', help='Start date for commit analysis (format: YYYY-MM-DD)')
    parser.add_argument('--end-date', help='End date for commit analysis (format: YYYY-MM-DD)')
    parser.add_argument('--limit-commits', type=int, help='Limit the number of commits to analyze')
    parser.add_argument('--analyze-local-commits', action='store_true', help='Analyze file sizes per commit from local repo (no checkout)')
    parser.add_argument('--local-path', default='.', help='Path to local git project (default: current directory)')
    parser.add_argument('--local-branch', help='Nama branch lokal yang ingin di-checkout sebelum analisis')
    parser.add_argument('--analyze-local-snapshot', action='store_true', help='Analyze all files in HEAD (snapshot)')
    parser.add_argument('--analyze-local-all-commits', action='store_true', help='Analyze all commits in local repo (not just linear)')
    parser.add_argument('--analyze-apk', action='store_true', help='Analyze APK/AAB file content')
    parser.add_argument('--apk-path', help='Path to APK/AAB file to analyze')
    parser.add_argument('--snapshot-file-types', help='Filter file types (comma separated, e.g. png,jpg,webp) for snapshot HEAD')
    args = parser.parse_args()

    # Validasi: jika TIDAK mode analisis lokal atau analyze-apk, argumen GitLab wajib
    if not (args.analyze_local_commits or args.analyze_local_snapshot or args.analyze_local_all_commits or args.analyze_apk):
        if not args.gitlab_url or not args.token or not args.project_id:
            parser.error("--gitlab-url, --token, dan --project-id wajib diisi kecuali menggunakan mode analisis lokal atau --analyze-apk")

    # Validasi argument
    if not args.analyze_branch and not args.target_branch and not (args.analyze_local_commits or args.analyze_local_snapshot or args.analyze_local_all_commits or args.analyze_apk):
        parser.error("Either --target-branch, --analyze-branch, atau salah satu mode analisis lokal/--analyze-apk harus diisi")

    # Validasi format tanggal
    if args.start_date:
        try:
            datetime.strptime(args.start_date, '%Y-%m-%d')
        except ValueError:
            parser.error("Start date must be in YYYY-MM-DD format")
    if args.end_date:
        try:
            datetime.strptime(args.end_date, '%Y-%m-%d')
        except ValueError:
            parser.error("End date must be in YYYY-MM-DD format")
    return args

def main():
    try:
        args = parse_args()
        # Analisis APK/AAB
        if getattr(args, 'analyze_apk', False):
            if not args.apk_path:
                print("❌ Harap isi --apk-path untuk analisis APK/AAB!")
                return
            project_root = args.local_path if hasattr(args, 'local_path') else None
            apk_df, mapping_df = analyze_apk_aab(args.apk_path, project_root)
            with pd.ExcelWriter(args.output_excel, engine='openpyxl') as writer:
                apk_df.to_excel(writer, sheet_name='APK_AAB_Content', index=False)
                if mapping_df is not None and not mapping_df.empty:
                    mapping_df.to_excel(writer, sheet_name='APK_to_Project_Mapping', index=False)
            print(f"\n✅ APK/AAB content report saved to {args.output_excel}")
            return
        # Hybrid mode: snapshot HEAD dan/atau all commits
        if getattr(args, 'analyze_local_snapshot', False) or getattr(args, 'analyze_local_all_commits', False):
            info_data = []
            from datetime import datetime
            info_data.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            info_data.append(["Repo Path", args.local_path])
            info_data.append(["Branch", get_current_branch(args.local_path)])
            if getattr(args, 'analyze_local_snapshot', False):
                info_data.append(["Snapshot HEAD", "Yes"])
            if getattr(args, 'analyze_local_all_commits', False):
                info_data.append(["All Commits", "Yes"])
            info_data.append(["Start Date", args.start_date if args.start_date else "(tidak dispesifikasikan)"])
            info_data.append(["End Date", args.end_date if args.end_date else "(tidak dispesifikasikan)"])
            file_types = args.snapshot_file_types.split(',') if getattr(args, 'snapshot_file_types', None) else None
            # Sheet: Snapshot HEAD
            snapshot_df, opt_df = analyze_local_snapshot(args.local_path, file_types=file_types) if getattr(args, 'analyze_local_snapshot', False) else (None, None)
            # Sheet: All Commits
            all_commits_df = analyze_local_all_commits(args.local_path, args.file_patterns.split(',') if args.file_patterns else None) if getattr(args, 'analyze_local_all_commits', False) else None
            # Sheet: Validation Rules
            validation_rules = [
                ["Kategori", "Ekstensi/Format", "Batas Maksimal (MB)", "Catatan"],
                ["Icon/Ilustrasi Sederhana", "XML (Vector)", "0.02", "< 20 KB"],
                ["Icon/Ilustrasi Sederhana", "PNG/JPG", "0.05", "≤ 50 KB"],
                ["Gambar Konten", "WebP", "0.2", "≤ 200 KB"],
                ["Gambar Fullscreen", "WebP/JPG", "0.5", "≤ 500 KB (1080x1920)"],
                ["Audio Efek", "OGG/AAC", "0.1", "< 100 KB (<5s)"],
                ["Audio Musik Pendek", "OGG/AAC", "0.3", "≤ 300 KB"],
                ["Video Pendek", "MP4/MOV/M4V", "1", "< 1 MB (480p)"],
                ["Lottie Animation", "JSON", "0.2", "50–200 KB"],
                ["Native Library", ".so", "5", "≤ 5 MB per ABI"],
                ["DEX/Kode", ".dex", "10", "≤ 10 MB per file"],
                ["JSON/Data Bundling", ".json", "0.1", "≤ 100 KB"],
                ["Font", ".ttf/.otf", "0.5", "≤ 500 KB"],
                ["Resource XML", ".xml", "0.02", "< 20 KB"],
            ]
            with pd.ExcelWriter(args.output_excel, engine='openpyxl') as writer:
                pd.DataFrame(info_data, columns=["Info", "Value"]).to_excel(writer, sheet_name='Info', index=False)
                if snapshot_df is not None:
                    snapshot_df.to_excel(writer, sheet_name='Snapshot HEAD', index=False)
                if opt_df is not None and not opt_df.empty:
                    opt_df.to_excel(writer, sheet_name='Optimization Candidates', index=False)
                if all_commits_df is not None:
                    all_commits_df.to_excel(writer, sheet_name='All Commits', index=False)
                pd.DataFrame(validation_rules[1:], columns=validation_rules[0]).to_excel(writer, sheet_name='Validation Rules', index=False)
            print(f"\n✅ Hybrid report saved to {args.output_excel}")
            return
        # Jalankan analisis lokal dulu, jika dipilih
        if getattr(args, 'analyze_local_commits', False):
            # Jika user ingin checkout branch tertentu
            if getattr(args, 'local_branch', None):
                if has_uncommitted_changes(args.local_path):
                    print("❌ Repo memiliki perubahan yang belum di-commit. Silakan commit atau stash dulu sebelum melanjutkan.")
                    return
                if not checkout_branch(args.local_path, args.local_branch):
                    return
            # Tampilkan branch yang sedang aktif
            current_branch = get_current_branch(args.local_path)
            print(f"\n🔎 Sedang menganalisis branch: {current_branch}\n")
            file_patterns = args.file_patterns.split(',') if args.file_patterns else None
            start_date = args.start_date
            end_date = args.end_date
            output_excel = args.output_excel if hasattr(args, 'output_excel') else 'local_commit_report.xlsx'
            analyze_local_commits(
                args.local_path,
                start_date=start_date,
                end_date=end_date,
                file_patterns=file_patterns,
                output_excel=output_excel
            )
            return

        # Inisialisasi GitLabAPI hanya jika mode lokal TIDAK dipilih
        gitlab = GitLabAPI(
            args.gitlab_url, 
            args.token, 
            args.project_id,
            verify_ssl=not args.no_verify_ssl
        )
        
        if args.analyze_branch:
            # Check branch and commits first
            print("\n🔍 Checking branch and commits...")
            if not gitlab.check_branch_and_commits(args.analyze_branch):
                print("❌ Branch check failed. Please verify the branch name and access.")
                return
                
            # Continue with existing analysis...
            file_analyzer = FileAnalyzer(
                file_patterns=args.file_patterns.split(',') if args.file_patterns else None,
                min_size_kb=args.min_size_kb,
                max_size_kb=args.max_size_kb
            )
            report_generator = ReportGenerator(args.output_excel)
            
            # Siapkan tanggal untuk filter
            start_date = None
            if args.start_date:
                start_date = datetime.strptime(args.start_date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
                print(f"📅 Start date: {start_date.strftime('%Y-%m-%d')}")
            
            end_date = None
            if args.end_date:
                end_date = datetime.strptime(args.end_date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
                print(f"📅 End date: {end_date.strftime('%Y-%m-%d')}")
            
            # Definisikan callback untuk memproses commit
            def process_commit(commit):
                # Filter berdasarkan title commit jika diinginkan
                if args.commit_title and args.commit_title.lower() not in commit['title'].lower():
                    return
                    
                changes = gitlab.get_commit_changes(commit['id'])
                commit_date = datetime.fromisoformat(commit['created_at'].replace('Z', '+00:00'))
                
                total_size = sum(change.get('size_kb', 0) or 0 for change in changes)
                print(f"\nProcessing commit {commit['id'][:8]} | "
                      f"files: {len(changes)} | "
                      f"size: {total_size:.2f}KB | "
                      f"date: {commit_date.strftime('%Y-%m-%d')}")
                
                for change in changes:
                    file_path = change.get("new_path", "unknown")
                    file_size = change.get('size_kb')
                    # Tampilkan semua file tanpa filter
                    report_generator.add_data(
                        commit['title'],
                        commit['id'][:8],
                        f"commit ({commit_date.strftime('%Y-%m-%d')})",
                        file_path,
                        file_size,
                        is_non_standard(file_path)
                    )
            
            # Mulai proses commit
            print("\n📦 Starting commit analysis...")
            gitlab.get_commits_by_branch(
                args.analyze_branch, 
                None,  # Tidak perlu branch_creation_date
                limit=args.limit_commits,
                start_date=start_date,
                end_date=end_date,
                callback=process_commit
            )
            print("\n✅ Commit analysis completed!")
            
        else:
            # Existing MR analysis code
            states = args.mr_state.split(',')
            mrs = gitlab.get_merge_requests(
                args.target_branch, 
                states,
                commit_message=args.commit_message
            )
            
            if not mrs:
                print("⚠️ No Merge Requests found.")
                return
            
            with tqdm(mrs, desc="📦 Processing", unit="mr", position=0, leave=True,
                     bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]') as pbar:
                for mr in pbar:
                    changes = gitlab.get_mr_changes(mr['iid'])
                    pbar.set_postfix_str(f"mr: #{mr['iid']} | files: {len(changes)}")
                    
                    for change in changes:
                        file_path = change.get("new_path", "unknown")
                        file_size = file_analyzer.get_file_size(file_path)
                        report_generator.add_data(
                            mr['title'],
                            mr['iid'],
                            mr['state'],
                            file_path,
                            file_size,
                            is_non_standard(file_path)
                        )
        
        # Generate report
        with tqdm(total=2, desc="📊 Generating", position=0, leave=True,
                 bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]') as pbar:
            df = report_generator.generate_excel()
            pbar.update(1)
            pbar.set_postfix_str("excel")
            
            # Generate plots if not disabled
            if not args.no_plot:
                report_generator.generate_plots(df)
                pbar.update(1)
                pbar.set_postfix_str("plots")
        
        # Print error summary at the end
        gitlab.error_collector.print_errors()
        file_analyzer.error_collector.print_errors()
        report_generator.error_collector.print_errors()
        
        print("✅ Analysis completed!")
        
    except KeyboardInterrupt:
        print(f"\n{Fore.YELLOW}⚠️ Proses dibatalkan oleh user{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}ℹ️ Data yang sudah diproses akan disimpan{Style.RESET_ALL}")
        sys.exit(0)
    except Exception as e:
        print(f"\n{Fore.RED}❌ Error: {str(e)}{Style.RESET_ALL}")
        sys.exit(1)

def analyze_local_commits(repo_path, start_date=None, end_date=None, file_patterns=None, output_excel="local_commit_report.xlsx"):
    import subprocess
    import pandas as pd
    import re
    from datetime import datetime
    from tqdm import tqdm
    # Get list of commits in date range
    log_cmd = [
        "git", "-C", repo_path, "log", "--pretty=format:%H|%ad|%s", "--date=short"
    ]
    if start_date:
        log_cmd.append(f"--since={start_date}")
    if end_date:
        log_cmd.append(f"--until={end_date}")
    commits = []
    result = subprocess.run(log_cmd, capture_output=True, text=True)
    for line in result.stdout.strip().splitlines():
        if '|' in line:
            parts = line.split('|', 2)
            if len(parts) == 3:
                sha, date, title = parts
                commits.append((sha.strip(), date.strip(), title.strip()))
            else:
                sha, date = parts[0], parts[1]
                commits.append((sha.strip(), date.strip(), ""))
    print(f"\n🔍 Found {len(commits)} commits in range.")
    data = []
    for sha, date, title in tqdm(commits, desc="Processing commits", unit="commit"):
        diff_cmd = [
            "git", "-C", repo_path, "diff-tree", "--no-commit-id", "--name-only", "-r", "-m", "--root", sha
        ]
        diff_result = subprocess.run(diff_cmd, capture_output=True, text=True)
        files = diff_result.stdout.strip().splitlines()
        for file_path in files:
            if file_patterns and not any(re.search(p, file_path) for p in file_patterns):
                continue
            size_kb = get_file_size_in_commit(repo_path, sha, file_path)
            size_mb = round(size_kb / 1024, 2) if size_kb is not None else None
            validation = validate_file_size(file_path, size_mb)
            data.append({
                "Commit": sha[:8],
                "Date": date,
                "Commit Title": title,
                "File": file_path,
                "File Size (MB)": size_mb,
                "NonStandard": is_non_standard(file_path),
                "Validation": validation
            })
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(["File Size (MB)"], ascending=[False])
    # Grouping by File
    if not df.empty:
        grouped = df.groupby("File").agg({
            "File Size (MB)": "max",
            "Commit": "count",
            "Date": "max",
            "Validation": lambda x: "OVERSIZE" if (x == "OVERSIZE").any() else "OK",
            "Commit Title": "last",
            "NonStandard": "first"
        }).reset_index()
        grouped = grouped.rename(columns={
            "Commit": "Change Count",
            "Date": "Last Change Date"
        })
    else:
        grouped = pd.DataFrame()
    # Summary info
    total_size = df["File Size (MB)"].sum() if not df.empty else 0
    total_files = len(df)
    total_oversize = (df["Validation"] == "OVERSIZE").sum() if not df.empty else 0
    biggest_file = df.iloc[0] if not df.empty else None
    info_data = [
        ["Branch", get_current_branch(repo_path)],
        ["Start Date", start_date if start_date else "(tidak dispesifikasikan)"],
        ["End Date", end_date if end_date else "(tidak dispesifikasikan)"],
        ["Total File Size (MB)", round(total_size, 2)],
        ["Total Files Analyzed", total_files],
        ["Total OVERSIZE Files", total_oversize],
    ]
    if biggest_file is not None:
        info_data += [
            ["Biggest File", biggest_file["File"]],
            ["Biggest File Size (MB)", biggest_file["File Size (MB)"]],
        ]
    # Tulis ke Excel dengan sheet info di tab pertama, lalu report, lalu aturan validasi
    validation_rules = [
        ["Kategori", "Ekstensi/Format", "Batas Maksimal (MB)", "Catatan"],
        ["Icon/Ilustrasi Sederhana", "XML (Vector)", "0.02", "< 20 KB"],
        ["Icon/Ilustrasi Sederhana", "PNG/JPG", "0.05", "≤ 50 KB"],
        ["Gambar Konten", "WebP", "0.2", "≤ 200 KB"],
        ["Gambar Fullscreen", "WebP/JPG", "0.5", "≤ 500 KB (1080x1920)"],
        ["Audio Efek", "OGG/AAC", "0.1", "< 100 KB (<5s)"],
        ["Audio Musik Pendek", "OGG/AAC", "0.3", "≤ 300 KB"],
        ["Video Pendek", "MP4/MOV/M4V", "1", "< 1 MB (480p)"],
        ["Lottie Animation", "JSON", "0.2", "50–200 KB"],
        ["Native Library", ".so", "5", "≤ 5 MB per ABI"],
        ["DEX/Kode", ".dex", "10", "≤ 10 MB per file"],
        ["JSON/Data Bundling", ".json", "0.1", "≤ 100 KB"],
        ["Font", ".ttf/.otf", "0.5", "≤ 500 KB"],
        ["Resource XML", ".xml", "0.02", "< 20 KB"],
    ]
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        pd.DataFrame(info_data, columns=["Info", "Value"]).to_excel(writer, sheet_name='Info', index=False)
        df.to_excel(writer, sheet_name='File Report', index=False)
        grouped.to_excel(writer, sheet_name='Grouped Files', index=False)
        pd.DataFrame(validation_rules[1:], columns=validation_rules[0]).to_excel(writer, sheet_name='Validation Rules', index=False)
    # Conditional formatting: warna merah untuk OVERSIZE
    try:
        from openpyxl.styles import PatternFill
        from openpyxl import load_workbook
        wb = load_workbook(output_excel)
        ws = wb['File Report']
        val_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Validation":
                val_col = idx
                break
        if val_col:
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=val_col, max_col=val_col):
                for cell in row:
                    if cell.value == "OVERSIZE":
                        cell.fill = red_fill
            wb.save(output_excel)
    except Exception as e:
        print(f"Warning: Conditional formatting failed: {e}")
    print(f"\n✅ Local commit file size report saved to {output_excel}")
    return df

def map_apk_to_project(apk_files, project_root):
    import os
    project_files = []
    for root, dirs, files in os.walk(project_root):
        for file in files:
            project_files.append(os.path.relpath(os.path.join(root, file), project_root))
    mapping = []
    for apk_file, size_mb, saran in apk_files:
        matches = [f for f in project_files if os.path.basename(f) == os.path.basename(apk_file)]
        mapping.append({
            'File in APK': apk_file,
            'Size (MB)': size_mb,
            'Project File': matches[0] if matches else '',
            'Saran Optimasi': saran
        })
    return mapping

def safe_rmtree(path, retries=3):
    for i in range(retries):
        try:
            shutil.rmtree(path)
            return
        except Exception as e:
            print(f"Retrying delete {path} ({i+1}/{retries})... {e}")
            time.sleep(1)
    raise

def analyze_apk_aab(apk_path, project_root=None):
    import os
    import pandas as pd
    from tqdm import tqdm
    extract_dir = apk_path + "_extract"
    if os.path.exists(extract_dir):
        safe_rmtree(extract_dir)
    with zipfile.ZipFile(apk_path, 'r') as zip_ref:
        zip_ref.extractall(extract_dir)
    data = []
    for root, dirs, files in os.walk(extract_dir):
        for file in files:
            file_path = os.path.join(root, file)
            rel_path = os.path.relpath(file_path, extract_dir)
            size_mb = round(os.path.getsize(file_path) / 1024 / 1024, 2)
            saran = suggest_optimization(rel_path, size_mb)
            data.append((rel_path, size_mb, saran))
    df = pd.DataFrame([{'File in APK/AAB': d[0], 'Size (MB)': d[1], 'Saran Optimasi': d[2]} for d in data])
    if not df.empty:
        df = df.sort_values(["Size (MB)"], ascending=[False])
    # Mapping ke project
    mapping_df = None
    if project_root:
        mapping = map_apk_to_project(data, project_root)
        mapping_df = pd.DataFrame(mapping)
        if not mapping_df.empty:
            mapping_df = mapping_df.sort_values(["Size (MB)"], ascending=[False])
    return df, mapping_df

if __name__ == '__main__':
    main()
